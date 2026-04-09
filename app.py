import io
import os
import re
from collections import OrderedDict
from datetime import date, datetime, time, timedelta
from html import unescape
from typing import Dict, Iterable, List, Tuple
from zoneinfo import ZoneInfo

import requests
from flask import Flask, Response, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = os.getenv("VIBE_BASE_URL", "https://vibecode.bitrix24.tech").rstrip("/")
API_KEY = os.getenv("VIBE_API_KEY", "")
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT_SECONDS", "60"))
APP_TIMEZONE = ZoneInfo(os.getenv("APP_TIMEZONE", "Asia/Vladivostok"))
PORTAL_BASE_URL = os.getenv("BITRIX24_PORTAL_URL", "https://iumiti.bitrix24.ru").rstrip("/")
PREVIEW_PAGE_SIZE = 20

OWNER_TYPES = OrderedDict(
    [
        ("lead", {"id": 1, "label": "Р›РёРґ"}),
        ("deal", {"id": 2, "label": "РЎРґРµР»РєР°"}),
    ]
)

KNOWN_TYPE_LABELS = {
    "CRM_TASKS_TASK": "Р—Р°РґР°С‡Р°",
    "CRM_TODO": "Р”РµР»Рѕ",
    "IMOPENLINES_SESSION": "Р§Р°С‚ РѕС‚РєСЂС‹С‚РѕР№ Р»РёРЅРёРё",
    "CALL": "Р—РІРѕРЅРѕРє",
    "EMAIL": "Email",
    "MEETING": "Р’СЃС‚СЂРµС‡Р°",
    "TASK": "Р—Р°РґР°С‡Р°",
    "NOTIFICATION": "РЈРІРµРґРѕРјР»РµРЅРёРµ",
}

TYPE_ID_LABELS = {
    1: "Р—РІРѕРЅРѕРє",
    2: "Р’СЃС‚СЂРµС‡Р°",
    3: "Р—Р°РґР°С‡Р°",
    4: "Email",
    5: "РЈРІРµРґРѕРјР»РµРЅРёРµ",
    6: "Р”РµР»Рѕ",
}

STATUS_OPTIONS = OrderedDict(
    [
        ("open", "РќРµ РІС‹РїРѕР»РЅРµРЅР°"),
        ("done", "Р’С‹РїРѕР»РЅРµРЅР°"),
    ]
)

app = Flask(__name__)


class ApiError(RuntimeError):
    pass


def api_request(method: str, path: str, *, params=None, json_body=None):
    if not API_KEY:
        raise ApiError("РќРµ Р·Р°РґР°РЅ VIBE_API_KEY")

    response = requests.request(
        method=method,
        url=f"{BASE_URL}{path}",
        headers={"X-Api-Key": API_KEY},
        params=params,
        json=json_body,
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("success"):
        raise ApiError(payload.get("error", "VibeCode API РІРµСЂРЅСѓР» РѕС€РёР±РєСѓ"))
    return payload


def parse_date_value(raw_value: str | None) -> date | None:
    if not raw_value:
        return None
    return date.fromisoformat(raw_value)


def to_iso_start(day: date | None) -> str | None:
    if not day:
        return None
    return datetime.combine(day, time.min, tzinfo=APP_TIMEZONE).isoformat()


def to_iso_end(day: date | None) -> str | None:
    if not day:
        return None
    end_of_day = datetime.combine(day + timedelta(days=1), time.min, tzinfo=APP_TIMEZONE) - timedelta(seconds=1)
    return end_of_day.isoformat()


def build_activity_filter(owner_type_id: int, date_from: date | None, date_to: date | None) -> dict:
    activity_filter = {"ownerTypeId": owner_type_id}
    created_at = {}
    if date_from:
        created_at["$gte"] = to_iso_start(date_from)
    if date_to:
        created_at["$lte"] = to_iso_end(date_to)
    if created_at:
        activity_filter["createdAt"] = created_at
    return activity_filter


def fetch_all_activities(owner_type_id: int, date_from: date | None, date_to: date | None) -> List[dict]:
    items: List[dict] = []
    page = 1
    page_size = 500
    while True:
        payload = api_request(
            "POST",
            "/v1/activities/search",
            json_body={
                "filter": build_activity_filter(owner_type_id, date_from, date_to),
                "page": page,
                "pageSize": page_size,
            },
        )
        data = payload.get("data", [])
        meta = payload.get("meta", {})
        items.extend(data)
        if not meta.get("hasMore"):
            break
        page += 1
    return items


def clean_html_text(value) -> str:
    if value in (None, ""):
        return ""
    text = str(value).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"[ \t]+", " ", unescape(text.replace("&nbsp;", " "))).strip()


def entity_label(owner_type_id: int) -> str:
    for item in OWNER_TYPES.values():
        if item["id"] == owner_type_id:
            return item["label"]
    return f"РЎСѓС‰РЅРѕСЃС‚СЊ {owner_type_id}"


def status_label(activity: dict) -> str:
    return STATUS_OPTIONS["done"] if activity.get("completed") else STATUS_OPTIONS["open"]


def normalized_type_key(activity: dict) -> str:
    provider_id = activity.get("PROVIDER_ID")
    if provider_id:
        return provider_id
    return f"TYPE_{activity.get('typeId') or activity.get('TYPE_ID') or 'unknown'}"


def type_label(activity: dict) -> str:
    provider_id = activity.get("PROVIDER_ID")
    if provider_id and provider_id in KNOWN_TYPE_LABELS:
        return KNOWN_TYPE_LABELS[provider_id]
    provider_type = activity.get("PROVIDER_TYPE_ID")
    if provider_type and provider_type in KNOWN_TYPE_LABELS:
        return KNOWN_TYPE_LABELS[provider_type]
    type_id = activity.get("typeId") or activity.get("TYPE_ID")
    if type_id in TYPE_ID_LABELS:
        return TYPE_ID_LABELS[type_id]
    if provider_id:
        return provider_id
    return f"РўРёРї {type_id}"


def normalize_entity_title(entity_name: str, item: dict) -> str:
    if entity_name == "deals":
        return item.get("title") or item.get("TITLE") or f"РЎРґРµР»РєР° #{item.get('id') or item.get('ID')}"
    if entity_name == "leads":
        parts = [
            item.get("TITLE"),
            item.get("NAME"),
            item.get("LAST_NAME"),
            item.get("COMPANY_TITLE"),
        ]
        title = " ".join(str(part).strip() for part in parts if part)
        return title or f"Р›РёРґ #{item.get('ID') or item.get('id')}"
    if entity_name == "users":
        parts = [
            item.get("name") or item.get("NAME"),
            item.get("lastName") or item.get("LAST_NAME"),
        ]
        full_name = " ".join(str(part).strip() for part in parts if part)
        return full_name or item.get("email") or item.get("EMAIL") or f"РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ #{item.get('id') or item.get('ID')}"
    return str(item.get("id") or item.get("ID") or "")


def chunked(values: Iterable[int], size: int) -> Iterable[List[int]]:
    chunk: List[int] = []
    for value in values:
        chunk.append(value)
        if len(chunk) == size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


def fetch_entity_map(entity_name: str, ids: Iterable[int]) -> Dict[int, str]:
    normalized_ids = sorted({int(item) for item in ids if item})
    if not normalized_ids:
        return {}

    result: Dict[int, str] = {}
    for group in chunked(normalized_ids, 50):
        calls = [
            {"id": f"{entity_name}-{entity_id}", "entity": entity_name, "action": "get", "entityId": entity_id}
            for entity_id in group
        ]
        payload = api_request("POST", "/v1/batch", json_body={"calls": calls})
        responses = payload.get("data", {}).get("results", {})
        for entity_id in group:
            item = responses.get(f"{entity_name}-{entity_id}")
            if isinstance(item, list):
                item = item[0] if item else None
            if item:
                result[entity_id] = normalize_entity_title(entity_name, item)
    return result


def enrich_activities(activities: List[dict]) -> List[dict]:
    deal_ids = [item.get("ownerId") for item in activities if item.get("ownerTypeId") == OWNER_TYPES["deal"]["id"]]
    lead_ids = [item.get("ownerId") for item in activities if item.get("ownerTypeId") == OWNER_TYPES["lead"]["id"]]
    user_ids = [item.get("responsibleId") for item in activities]

    deal_titles = fetch_entity_map("deals", deal_ids)
    lead_titles = fetch_entity_map("leads", lead_ids)
    user_titles = fetch_entity_map("users", user_ids)

    enriched = []
    for activity in activities:
        owner_type_id = activity.get("ownerTypeId")
        owner_id = activity.get("ownerId")
        owner_title = ""
        if owner_type_id == OWNER_TYPES["deal"]["id"]:
            owner_title = deal_titles.get(owner_id, f"РЎРґРµР»РєР° #{owner_id}")
        elif owner_type_id == OWNER_TYPES["lead"]["id"]:
            owner_title = lead_titles.get(owner_id, f"Р›РёРґ #{owner_id}")

        item = dict(activity)
        item["ownerTypeLabel"] = entity_label(owner_type_id)
        item["ownerTitle"] = owner_title
        item["typeKey"] = normalized_type_key(activity)
        item["typeLabel"] = type_label(activity)
        item["statusKey"] = "done" if activity.get("completed") else "open"
        item["statusLabel"] = status_label(activity)
        item["responsibleName"] = user_titles.get(activity.get("responsibleId"), str(activity.get("responsibleId") or ""))
        enriched.append(item)

    return enriched


def filter_activities(activities: List[dict], selected_types: List[str], selected_statuses: List[str]) -> List[dict]:
    status_set = set(selected_statuses)
    type_set = set(selected_types)
    result = []
    for activity in activities:
        if type_set and activity["typeKey"] not in type_set:
            continue
        if status_set and activity["statusKey"] not in status_set:
            continue
        result.append(activity)
    return result


def build_filter_options(activities: List[dict]) -> Tuple[List[dict], List[dict]]:
    types = OrderedDict()
    for activity in activities:
        types.setdefault(activity["typeKey"], activity["typeLabel"])

    type_options = [{"value": key, "label": label} for key, label in types.items()]
    status_options = [{"value": key, "label": label} for key, label in STATUS_OPTIONS.items()]
    return type_options, status_options


def format_preview_datetime(value) -> str:
    if not value:
        return ""
    raw_value = str(value)
    try:
        dt = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone(APP_TIMEZONE)
        return dt.strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return raw_value


def entity_detail_url(owner_type_id: int, owner_id: int | None) -> str:
    if not owner_id:
        return ""
    if owner_type_id == OWNER_TYPES["deal"]["id"]:
        return f"{PORTAL_BASE_URL}/crm/deal/details/{owner_id}/"
    if owner_type_id == OWNER_TYPES["lead"]["id"]:
        return f"{PORTAL_BASE_URL}/crm/lead/details/{owner_id}/"
    return ""


def serialize_preview_rows(activities: List[dict], limit: int | None = None) -> List[dict]:
    rows = []
    source = activities if limit is None else activities[:limit]
    for activity in source:
        rows.append(
            {
                "id": activity.get("id"),
                "entityId": activity.get("ownerId"),
                "entity": activity.get("ownerTypeLabel"),
                "entityTitle": activity.get("ownerTitle"),
                "type": activity.get("typeLabel"),
                "status": activity.get("statusLabel"),
                "subject": clean_html_text(activity.get("subject")),
                "createdAt": format_preview_datetime(activity.get("createdAt")),
                "deadline": format_preview_datetime(activity.get("deadline")),
                "responsible": activity.get("responsibleName"),
                "entityUrl": entity_detail_url(activity.get("ownerTypeId"), activity.get("ownerId")),
            }
        )
    return rows


def sort_preview_rows(rows: List[dict], sort_by: str, sort_dir: str) -> List[dict]:
    allowed = {
        "id": "entityId",
        "entity": "entity",
        "title": "entityTitle",
        "type": "type",
        "status": "status",
        "subject": "subject",
        "createdAt": "createdAt",
        "deadline": "deadline",
        "responsible": "responsible",
    }
    sort_key = allowed.get(sort_by, "createdAt")
    reverse = sort_dir == "desc"

    def key_func(item: dict):
        value = item.get(sort_key, "")
        if sort_key == "entityId":
            return int(value or 0)
        return str(value or "").lower()

    return sorted(rows, key=key_func, reverse=reverse)


def paginate_rows(rows: List[dict], page: int, page_size: int = PREVIEW_PAGE_SIZE) -> tuple[list[dict], int]:
    total_pages = max(1, (len(rows) + page_size - 1) // page_size)
    current_page = min(max(page, 1), total_pages)
    start = (current_page - 1) * page_size
    end = start + page_size
    return rows[start:end], total_pages


def collect_activities(owner_keys: List[str], date_from: date | None, date_to: date | None) -> List[dict]:
    selected_owners = owner_keys or list(OWNER_TYPES.keys())
    activities: List[dict] = []
    for owner_key in selected_owners:
        owner_type_id = OWNER_TYPES[owner_key]["id"]
        activities.extend(fetch_all_activities(owner_type_id, date_from, date_to))

    activities.sort(key=lambda item: item.get("createdAt") or "", reverse=True)
    return enrich_activities(activities)


def create_workbook(activities: List[dict]) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "РђРєС‚РёРІРЅРѕСЃС‚Рё"

    headers = [
        "ID Р°РєС‚РёРІРЅРѕСЃС‚Рё",
        "РЎСѓС‰РЅРѕСЃС‚СЊ CRM",
        "ID СЃСѓС‰РЅРѕСЃС‚Рё",
        "РќР°Р·РІР°РЅРёРµ СЃСѓС‰РЅРѕСЃС‚Рё",
        "РўРёРї Р°РєС‚РёРІРЅРѕСЃС‚Рё",
        "РЎС‚Р°С‚СѓСЃ Р°РєС‚РёРІРЅРѕСЃС‚Рё",
        "РўРµРјР°",
        "Р”Р°С‚Р° СЃРѕР·РґР°РЅРёСЏ",
        "Р”Р°С‚Р° Р·Р°РІРµСЂС€РµРЅРёСЏ",
        "РљСЂР°Р№РЅРёР№ СЃСЂРѕРє",
        "РћС‚РІРµС‚СЃС‚РІРµРЅРЅС‹Р№",
        "РџСЂРѕРІР°Р№РґРµСЂ",
        "РћРїРёСЃР°РЅРёРµ",
    ]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F766E")

    for activity in activities:
        worksheet.append(
            [
                activity.get("id"),
                activity.get("ownerTypeLabel"),
                activity.get("ownerId"),
                activity.get("ownerTitle"),
                activity.get("typeLabel"),
                activity.get("statusLabel"),
                clean_html_text(activity.get("subject")),
                activity.get("createdAt"),
                activity.get("endTime"),
                activity.get("deadline"),
                activity.get("responsibleName"),
                activity.get("PROVIDER_ID") or "",
                clean_html_text(activity.get("description")),
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 14), 48)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/")
def index():
    today = date.today()
    default_from = today - timedelta(days=30)
    owner_keys = list(OWNER_TYPES.keys())
    activities = collect_activities(owner_keys, default_from, today)
    type_options, status_options = build_filter_options(activities)
    preview_rows = serialize_preview_rows(activities, PREVIEW_PAGE_SIZE)

    return render_template(
        "index.html",
        owner_types=OWNER_TYPES,
        default_from=default_from.isoformat(),
        default_to=today.isoformat(),
        type_options=type_options,
        status_options=status_options,
        initial_count=len(activities),
        preview_rows=preview_rows,
        preview_limit=len(preview_rows),
    )


@app.get("/api/filter-options")
def filter_options():
    owner_keys = request.args.getlist("owner")
    date_from = parse_date_value(request.args.get("date_from"))
    date_to = parse_date_value(request.args.get("date_to"))
    activities = collect_activities(owner_keys, date_from, date_to)
    type_options, status_options = build_filter_options(activities)
    return jsonify(
        {
            "types": type_options,
            "statuses": status_options,
            "count": len(activities),
        }
    )


@app.get("/api/preview")
def preview():
    owner_keys = request.args.getlist("owner")
    date_from = parse_date_value(request.args.get("date_from"))
    date_to = parse_date_value(request.args.get("date_to"))
    selected_types = request.args.getlist("activity_type")
    selected_statuses = request.args.getlist("activity_status")
    sort_by = request.args.get("sort_by", "createdAt")
    sort_dir = request.args.get("sort_dir", "desc")
    page = int(request.args.get("page", "1"))

    activities = collect_activities(owner_keys, date_from, date_to)
    filtered = filter_activities(activities, selected_types, selected_statuses)
    rows = serialize_preview_rows(filtered, None)
    rows = sort_preview_rows(rows, sort_by, sort_dir)
    paged_rows, total_pages = paginate_rows(rows, page)
    return jsonify(
        {
            "count": len(filtered),
            "rows": paged_rows,
            "page": min(max(page, 1), total_pages),
            "pageSize": PREVIEW_PAGE_SIZE,
            "totalPages": total_pages,
            "sortBy": sort_by,
            "sortDir": sort_dir,
        }
    )


@app.post("/export")
def export():
    owner_keys = request.form.getlist("owner")
    date_from = parse_date_value(request.form.get("date_from"))
    date_to = parse_date_value(request.form.get("date_to"))
    selected_types = request.form.getlist("activity_type")
    selected_statuses = request.form.getlist("activity_status")

    activities = collect_activities(owner_keys, date_from, date_to)
    filtered = filter_activities(activities, selected_types, selected_statuses)
    workbook_buffer = create_workbook(filtered)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        workbook_buffer,
        as_attachment=True,
        download_name=f"crm_activities_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.errorhandler(Exception)
def handle_error(error):
    status_code = 500
    message = str(error)
    if isinstance(error, requests.HTTPError) and error.response is not None:
        status_code = error.response.status_code
        try:
            message = error.response.json()
        except ValueError:
            message = error.response.text
    if request.path.startswith("/api/"):
        return jsonify({"error": message}), status_code
    return Response(f"РћС€РёР±РєР° РїСЂРёР»РѕР¶РµРЅРёСЏ: {message}", status=status_code, mimetype="text/plain; charset=utf-8")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000, debug=False)
